"""
源数据限定导入脚本
用途：将两张源Excel一次性导入数据库，测试完成后正式用户改用页面批量维护。

限定1（特殊款式限定位置）：4列 - 款式 | 位置 | 印花 | 限定款式
  形式1：A=款式(合并单元格可延续), B=位置, C=印花(空=不限), D=空
        → rule_type='style_position'
  形式2：A=空, B=单位置, C=多印花, D=多款式(逗号分隔)
        → rule_type='position_print'
  形式3：A=款式, B/C/D 全空
        → rule_type='style_ban'

限定2（魔术贴限定）：3列 - 印花编码 | 多款式 | 多位置
  展开为每个款式 × 每个位置，该印花只能贴这些组合
  → rule_type='print_restriction'

用法：
  cd C:/Users/Administrator/PythonProjects/style-print-manager/backend
  python import_restrictions_source.py \
      --file1 "C:/Users/Administrator/Desktop/特殊款式限定位置.xlsx" \
      --file2 "C:/Users/Administrator/Desktop/魔术贴限定表.xlsx"

可选参数：
  --dry-run   只解析不写库，打印统计信息
  --file1     限定1文件路径（特殊款式限定位置）
  --file2     限定2文件路径（魔术贴限定）
"""

import sys
import re
import argparse
from pathlib import Path

# 让 backend 目录下的 app 包可被直接 import
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from app.database import SessionLocal
from app import crud, models

# ── 解析工具 ──────────────────────────────────────────────────

_SPECIAL_PRINTS = {"纯色", "福袋", "自搭"}


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _split_multi(s: str):
    return [x.strip() for x in re.split(r"[,，、\n\r]+", s) if x.strip()]


def _parse_print_list(raw: str):
    """返回 None = 不限；返回 list = 仅允许列表中印花（已去重排序）"""
    if not raw:
        return None
    codes = [c for c in _split_multi(raw) if c not in _SPECIAL_PRINTS]
    return sorted(set(codes)) if codes else None


# ── 写库操作 ────────────────────────────────────────────────

def _upsert_style_position_rule(db, style_code: str, position_name: str, new_prints, dry_run: bool):
    """形式1：款式+位置 → 印花白名单"""
    style = crud.get_style_by_code(db, style_code)
    if not style:
        raise ValueError(f"款式 '{style_code}' 不在款式表中")
    position = crud.get_position_by_name(db, position_name)
    if not position:
        raise ValueError(f"位置 '{position_name}' 不在位置表中")

    if dry_run:
        return

    existing = db.query(models.StylePositionRule).filter(
        models.StylePositionRule.rule_type == 'style_position',
        models.StylePositionRule.style_id == style.id,
        models.StylePositionRule.position_id == position.id
    ).first()

    if existing:
        if new_prints is None:
            # 新规则说"不限" → 覆盖为 NULL
            existing.allowed_prints = None
        else:
            # 合并印花白名单
            if existing.allowed_prints:
                old_set = set(existing.allowed_prints.split(","))
                merged = sorted(old_set | set(new_prints))
                existing.allowed_prints = ",".join(merged)
            else:
                existing.allowed_prints = ",".join(new_prints)
        db.commit()
    else:
        ap = ",".join(new_prints) if new_prints else None
        db.add(models.StylePositionRule(
            rule_type='style_position',
            style_id=style.id,
            position_id=position.id,
            allowed_prints=ap,
        ))
        db.commit()


def _upsert_position_print_rule(db, position_name: str, print_code: str, style_ids: list[int], dry_run: bool):
    """形式2：位置+印花 → 款式白名单"""
    position = crud.get_position_by_name(db, position_name)
    if not position:
        raise ValueError(f"位置 '{position_name}' 不在位置表中")

    if not crud.get_print_by_code(db, print_code):
        raise ValueError(f"印花 '{print_code}' 不在印花表中")

    if dry_run:
        return

    existing = db.query(models.StylePositionRule).filter(
        models.StylePositionRule.rule_type == 'position_print',
        models.StylePositionRule.position_id == position.id,
        models.StylePositionRule.print_code == print_code
    ).first()

    if existing:
        # 合并款式白名单
        old_ids = set(existing.allowed_styles.split(','))
        new_ids = old_ids | set(map(str, style_ids))
        existing.allowed_styles = ','.join(sorted(new_ids, key=int))
        db.commit()
    else:
        db.add(models.StylePositionRule(
            rule_type='position_print',
            position_id=position.id,
            print_code=print_code,
            allowed_styles=','.join(map(str, style_ids))
        ))
        db.commit()


def _upsert_print_restriction_rule(db, print_code: str, combos: list[tuple[int, int]], dry_run: bool):
    """魔术贴限定：印花 → (款式,位置)组合白名单"""
    if not crud.get_print_by_code(db, print_code):
        raise ValueError(f"印花 '{print_code}' 不在印花表中")

    if dry_run:
        return

    existing = db.query(models.StylePositionRule).filter(
        models.StylePositionRule.rule_type == 'print_restriction',
        models.StylePositionRule.print_code == print_code
    ).first()

    combo_strs = [f"{sid}:{pid}" for sid, pid in combos]

    if existing:
        # 合并款式位置组合
        old_combos = set(existing.allowed_style_positions.split(','))
        new_combos = old_combos | set(combo_strs)
        existing.allowed_style_positions = ','.join(sorted(new_combos))
        db.commit()
    else:
        db.add(models.StylePositionRule(
            rule_type='print_restriction',
            print_code=print_code,
            allowed_style_positions=','.join(combo_strs)
        ))
        db.commit()


def _upsert_ban(db, style_code: str, dry_run: bool):
    """形式3：款式全禁"""
    style = crud.get_style_by_code(db, style_code)
    if not style:
        raise ValueError(f"款式 '{style_code}' 不在款式表中")
    if dry_run:
        return

    existing = db.query(models.StylePositionRule).filter(
        models.StylePositionRule.rule_type == 'style_ban',
        models.StylePositionRule.style_id == style.id
    ).first()

    if not existing:
        db.add(models.StylePositionRule(
            rule_type='style_ban',
            style_id=style.id
        ))
        db.commit()


# ── 限定1 解析逻辑 ────────────────────────────────────────────

def import_file1(path: str, db, dry_run: bool):
    """
    4列：款式 | 位置 | 印花 | 限定款式
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    rule_count, ban_count, form2_count, errors = 0, 0, 0, []
    last_style = None

    for row_idx, row in enumerate(rows[1:], 2):
        A = _str(row[0]) if len(row) > 0 else None
        B = _str(row[1]) if len(row) > 1 else None
        C = _str(row[2]) if len(row) > 2 else None
        D = _str(row[3]) if len(row) > 3 else None

        if not A and not B and not C and not D:
            continue

        try:
            # 形式3：款式全禁
            if A and not B and not C and not D:
                _upsert_ban(db, A, dry_run)
                ban_count += 1
                last_style = None
                continue

            # 形式2：位置+印花 → 多款式
            if not A and B and C and D:
                print_codes = _split_multi(C)
                style_codes = _split_multi(D)

                # 获取款式ID列表
                style_ids = []
                for sc in style_codes:
                    style = crud.get_style_by_code(db, sc)
                    if not style:
                        raise ValueError(f"款式 {sc} 不存在")
                    style_ids.append(style.id)

                for print_code in print_codes:
                    _upsert_position_print_rule(db, B, print_code, style_ids, dry_run)
                    form2_count += 1

                continue

            # 形式1：款式+位置 → 印花列表
            if A:
                last_style = A
            if last_style and B:
                print_list = _parse_print_list(C)
                _upsert_style_position_rule(db, last_style, B, print_list, dry_run)
                rule_count += 1

        except Exception as e:
            errors.append(f"第{row_idx}行: {e}")

    return rule_count, form2_count, ban_count, errors


# ── 限定2 解析逻辑 ────────────────────────────────────────────

def import_file2(path: str, db, dry_run: bool):
    """
    3列：印花编码 | 多款式（逗号分隔）| 多位置（逗号分隔）
    展开为：每个款式 × 每个位置 → 该印花只能贴这些组合
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    rule_count, errors = 0, []

    for row_idx, row in enumerate(rows[1:], 2):
        A = _str(row[0]) if len(row) > 0 else None  # 印花编码
        B = _str(row[1]) if len(row) > 1 else None  # 款式列表
        C = _str(row[2]) if len(row) > 2 else None  # 位置列表

        if not A or not B or not C:
            continue

        try:
            style_codes = _split_multi(B)
            position_names = _split_multi(C)

            # 构建款式位置组合
            combos = []
            for sc in style_codes:
                style = crud.get_style_by_code(db, sc)
                if not style:
                    raise ValueError(f"款式 {sc} 不存在")

                for pn in position_names:
                    position = crud.get_position_by_name(db, pn)
                    if not position:
                        raise ValueError(f"位置 {pn} 不存在")

                    combos.append((style.id, position.id))

            _upsert_print_restriction_rule(db, A, combos, dry_run)
            rule_count += 1

        except Exception as e:
            errors.append(f"第{row_idx}行: {e}")

    return rule_count, errors


# ── 主入口 ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="限定源数据导入脚本")
    parser.add_argument("--file1", default="C:/Users/Administrator/Desktop/特殊款式限定位置.xlsx",
                        help="限定1文件路径（特殊款式限定位置）")
    parser.add_argument("--file2", default="C:/Users/Administrator/Desktop/魔术贴限定表.xlsx",
                        help="限定2文件路径（魔术贴限定）")
    parser.add_argument("--dry-run", action="store_true",
                        help="只解析不写库，用于验证数据")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        mode = "[DRY RUN]" if args.dry_run else "[WRITE]"

        # ── 限定1 ──
        if Path(args.file1).exists():
            print(f"\n{mode} 导入限定1：{args.file1}")
            r1, r2, bans, errs = import_file1(args.file1, db, args.dry_run)
            print(f"  形式1（款式+位置→印花）：{r1} 条")
            print(f"  形式2（位置+印花→款式）：{r2} 条")
            print(f"  全禁款式：{bans} 条")
            if errs:
                print(f"  错误 ({len(errs)} 条)：")
                for e in errs[:20]:
                    print(f"    !! {e}")
                if len(errs) > 20:
                    print(f"    ...（共 {len(errs)} 条，只显示前 20）")
            else:
                print("  无错误")
        else:
            print(f"\n限定1文件不存在，跳过：{args.file1}")

        # ── 限定2 ──
        if Path(args.file2).exists():
            print(f"\n{mode} 导入限定2：{args.file2}")
            r, errs = import_file2(args.file2, db, args.dry_run)
            print(f"  印花限定规则：{r} 条")
            if errs:
                print(f"  错误 ({len(errs)} 条)：")
                for e in errs[:20]:
                    print(f"    !! {e}")
                if len(errs) > 20:
                    print(f"    ...（共 {len(errs)} 条，只显示前 20）")
            else:
                print("  无错误")
        else:
            print(f"\n限定2文件不存在，跳过：{args.file2}")

        print("\n完成。" if not args.dry_run else "\nDRY RUN 完成，未写入数据库。")
    finally:
        db.close()


if __name__ == "__main__":
    main()
