import sys
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import models, schemas
from app.crud import create_style_position_rule
from app.database import Base
from app.routers.excel_io import _import_positions_dict, _import_restrictions
from app.routers.excel_io import _import_prints, _import_styles


def make_db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    return SessionLocal()


def test_restrictions_with_missing_reference_do_not_replace_existing_rules():
    db = make_db()
    try:
        style = models.Style(code="S001")
        position = models.Position(code="P001", name="胸标", area="小图位置")
        print_obj = models.Print(code="PR001", name="现有印花")
        db.add_all([style, position, print_obj])
        db.commit()

        create_style_position_rule(
            db,
            schemas.StylePositionRuleCreate(
                rule_type=3,
                position_code="P001",
                style_codes=["S001"],
                print_codes=["PR001"],
            ),
        )

        wb = Workbook()
        ws = wb.active
        ws.append(["款式", "位置", "印花", "限定款式"])
        ws.append(["S001", "胸标", "MISSING_PRINT", None])

        count, errors, _, _, warnings = _import_restrictions(ws, db)

        assert count == 0
        assert any("没有可导入的有效限定规则" in error for error in errors)
        assert any("MISSING_PRINT" in warning for warning in warnings)
        rules = db.query(models.StylePositionRule).all()
        assert len(rules) == 1
        assert rules[0].print_ids == str(print_obj.id)
    finally:
        db.close()


def test_position_import_does_not_update_existing_code_name_binding():
    db = make_db()
    try:
        db.add(models.Position(code="P001", name="胸标", area="小图位置"))
        db.commit()

        count, errors, stats = _import_positions_dict(
            {"big_position": {"大图": "P001", "背标": "P002"}},
            db,
        )

        assert count == 1
        assert stats["created"] == 1
        assert any("位置编号 P001 已绑定名称 '胸标'" in error for error in errors)
        assert db.query(models.Position).filter_by(code="P001").one().name == "胸标"
        assert db.query(models.Position).filter_by(code="P002").one().name == "背标"
    finally:
        db.close()


def test_create_position_restriction_merges_existing_position_rule():
    db = make_db()
    try:
        styles = [models.Style(code=f"S00{i}") for i in range(1, 4)]
        prints = [models.Print(code=f"PR00{i}", name=f"印花{i}") for i in range(1, 3)]
        position = models.Position(code="P001", name="右下B标", area="小图位置")
        db.add_all([*styles, *prints, position])
        db.commit()

        first = create_style_position_rule(
            db,
            schemas.StylePositionRuleCreate(
                rule_type=2,
                position_code="P001",
                style_codes=["S001"],
                print_codes=["PR001"],
            ),
        )
        second = create_style_position_rule(
            db,
            schemas.StylePositionRuleCreate(
                rule_type=2,
                position_code="P001",
                style_codes=["S002"],
                print_codes=["PR002"],
            ),
        )
        third = create_style_position_rule(
            db,
            schemas.StylePositionRuleCreate(
                rule_type=2,
                position_code="P001",
                style_codes=["S003"],
                print_codes=None,
            ),
        )

        rules = db.query(models.StylePositionRule).filter_by(rule_type=2, position_id=position.id).all()
        assert len(rules) == 1
        assert first.id == second.id == third.id == rules[0].id
        assert rules[0].style_ids == ",".join(str(style.id) for style in styles)
        assert rules[0].print_ids == ",".join(str(print_obj.id) for print_obj in prints)
    finally:
        db.close()


def test_style_import_keeps_row_when_optional_numbers_are_invalid():
    db = make_db()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["白坯款式编码", "年份", "白坯重量", "吊牌价", "高价品牌吊牌价"])
        ws.append(["通DK桑蚕丝珠地POLO短T", "针织", "bad", "199", "bad"])

        count, errors, stats = _import_styles(ws, db)

        style = db.query(models.Style).filter_by(code="通DK桑蚕丝珠地POLO短T").one()
        assert count == 1
        assert errors == []
        assert stats["created"] == 1
        assert style.year is None
        assert style.blank_weight is None
        assert style.tag_price == 199
        assert style.premium_tag_price is None
        assert any("年份不是有效数字" in warning for warning in stats["warnings"])
    finally:
        db.close()


def test_print_import_supports_extended_headers_with_duplicate_replacement_columns():
    db = make_db()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append([
            '图案名称\n(小写字母,可重,不带"标"字）',
            "图案大小",
            "图案规格",
            "工艺属性",
            "商品编码",
            "真维斯款号\n（审核通过）",
            "替换的印花编码",
            "替换印花款号\n（公式下拉）",
            "JWCO款号\n（审核通过）",
            "替换的印花编码",
            "替换印花款号\n（公式下拉）",
            "city款号\n（审核通过）",
            "替换的印花编码",
            "替换印花款号\n（公式下拉）",
            "唐狮款号\n（审核通过）",
            "备注",
        ])
        ws.append([
            "ge紫色花环",
            "小图",
            "X",
            "高弹胶水微边烫画",
            "ge紫色花环X",
            "ZWX-STYLE",
            "ZWX-REPLACE-CODE",
            "ZWX-REPLACE-STYLE",
            "JWCO-STYLE",
            "JWCO-REPLACE-CODE",
            "JWCO-REPLACE-STYLE",
            "CITY-STYLE",
            "CITY-REPLACE-CODE",
            "CITY-REPLACE-STYLE",
            "TANGSHI-STYLE",
            "测试备注",
        ])

        count, errors, stats = _import_prints(ws, db)

        print_obj = db.query(models.Print).filter_by(code="ge紫色花环X").one()
        assert count == 1
        assert errors == []
        assert stats["created"] == 1
        assert print_obj.name == "ge紫色花环"
        assert print_obj.zwx_replace_code == "ZWX-REPLACE-CODE"
        assert print_obj.jwco_replace_code == "JWCO-REPLACE-CODE"
        assert print_obj.city_replace_code == "CITY-REPLACE-CODE"
        assert print_obj.tangshi_style_code == "TANGSHI-STYLE"
    finally:
        db.close()


def test_restriction_import_supports_special_position_headers():
    db = make_db()
    try:
        db.add_all([
            models.Style(code="女辣妹鱼骨短T"),
            models.Style(code="女不规则短款短T"),
            models.Position(code="P001", name="胸标", area="小图位置"),
            models.Position(code="P002", name="中标", area="小图位置"),
            models.Print(code="PRINT001", name="测试印花"),
        ])
        db.commit()

        wb = Workbook()
        ws = wb.active
        ws.append([
            "款式",
            "可贴图位置\n（位置名称需为标准名称）",
            "可贴的印花编码\n（需与印花基础信息表的印花编码一致,中间用,链接）",
            "限定款式\n（需跟产品信息综合表名称一致）",
        ])
        ws.append(["女辣妹鱼骨短T", "胸标", None, None])
        ws.append([None, "中标", None, None])
        ws.append([None, "胸标", "PRINT001", "女不规则短款短T"])
        ws.merge_cells("A2:A3")

        count, errors, filtered_count, skipped_duplicate_count, warnings = _import_restrictions(ws, db)

        rules = db.query(models.StylePositionRule).order_by(models.StylePositionRule.rule_type).all()
        assert count == 3
        assert errors == []
        assert warnings == []
        assert filtered_count == 0
        assert skipped_duplicate_count == 0
        assert [rule.rule_type for rule in rules] == [2, 3, 3]
    finally:
        db.close()


def test_restriction_import_ignores_missing_prints_in_lists():
    db = make_db()
    try:
        db.add_all([
            models.Style(code="S001"),
            models.Position(code="P001", name="胸标", area="小图位置"),
            models.Print(code="PRINT001", name="测试印花"),
        ])
        db.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["款式", "位置", "印花", "限定款式"])
        ws.append(["S001", "胸标", "PRINT001,MISSING_PRINT", None])

        count, errors, filtered_count, skipped_duplicate_count, warnings = _import_restrictions(ws, db)

        rule = db.query(models.StylePositionRule).one()
        assert count == 1
        assert errors == []
        assert filtered_count == 0
        assert skipped_duplicate_count == 0
        assert rule.print_ids == str(db.query(models.Print).filter_by(code="PRINT001").one().id)
        assert any("MISSING_PRINT" in warning for warning in warnings)
    finally:
        db.close()
