import io
import json
import datetime
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from .. import crud, schemas, models
from ..database import get_db

router = APIRouter(prefix="/api/excel", tags=["Excel导入导出"])

# 款式表 31 列（与原始模版保持一致）
STYLE_HEADERS = [
    "品牌属性", "属性", "面料种类", "年份", "性别", "季节",
    "类目", "商品分类", "虚拟分类",
    "商品款号", "白坯款式编码*",
    "款式备注",
    "在售颜色", "淘汰颜色", "颜色备注",
    "尺码", "号型", "尺码备注",
    "可印花范围",
    "面料成分", "Fabric Ingredients", "热风成分",
    "面料名称", "面料克重", "白坯重量",
    "开发时间", "吊牌价", "高价品牌吊牌价",
    "执行标准", "安全技术类别", "产品分类",
]
PRINT_HEADERS = [
    "图案名称*", "图案大小", "图案规格", "工艺属性", "商品编码*",
    "真维斯款号", "真维斯替换编码", "真维斯替换款号",
    "JWCO款号", "JWCO替换编码", "JWCO替换款号",
    "CITY款号", "CITY替换编码", "CITY替换款号",
    "唐狮款号", "备注",
]
POSITION_HEADERS = ["位置编号*", "位置名称*", "区域", "备注"]
RESTRICTION_HEADERS = ["白坯款式编码*", "位置名称*", "允许印花编码(逗号分隔，留空=不限)", "是否启用(1/0)", "备注"]


# ── 工具函数 ──────────────────────────────────────────

def _excel_date(val):
    """Excel 序列号 / datetime → Python date"""
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val if isinstance(val, datetime.date) else val.date()
    try:
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(val))
    except Exception:
        return None


def _str(v):
    if v is None:
        return None
    # 处理 pandas 的 NaN 值
    if pd.isna(v):
        return None
    s = str(v).strip()
    return s if s else None


def _header_key(v):
    s = _str(v)
    return s.rstrip("*").strip() if s else None


def _header_index(headers, *names):
    wanted = {name.rstrip("*").strip() for name in names}
    for idx, header in enumerate(headers):
        if _header_key(header) in wanted:
            return idx
    return None


def _require_headers(headers, required: dict[str, tuple[str, ...]]):
    missing = [label for label, names in required.items() if _header_index(headers, *names) is None]
    if missing:
        raise HTTPException(400, f"导入文件缺少关键列：{', '.join(missing)}")


def _apply_header_style(ws):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 20


def _parse_wb(content: bytes, filename: str) -> object:
    try:
        return load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"文件解析失败: {e}")


def _stream_wb(wb: object, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 核心导入逻辑（各模块独立） ────────────────────────

def _import_styles(ws, db: Session):
    """从 worksheet 导入款式，返回 (count, errors)"""
    headers = [cell.value for cell in ws[1]]
    _require_headers(headers, {"白坯款式编码": ("白坯款式编码",)})

    def col(row_vals, name):
        idx = _header_index(headers, name)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    count, errors = 0, []
    stats = {"created": 0, "updated": 0, "skipped": 0, "warnings": []}
    seen_codes: dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        code_val = col(row, "白坯款式编码")
        if not code_val:
            stats["skipped"] += 1
            continue
        code = str(code_val).strip()
        if not code:
            errors.append(f"第{row_idx}行：白坯款式编码不能为空")
            continue
        if code in seen_codes:
            stats["warnings"].append(f"第{row_idx}行：款式编码 {code} 与第{seen_codes[code]}行重复，后者覆盖前者")
        seen_codes[code] = row_idx
        try:
            payload = schemas.StyleCreate(
                code=code,
                product_code=_str(col(row, "商品款号")),
                brand_attr=_str(col(row, "品牌属性")),
                attr=_str(col(row, "属性")),
                fabric_type=_str(col(row, "面料种类")),
                year=int(col(row, "年份")) if col(row, "年份") is not None else None,
                gender=_str(col(row, "性别")),
                season=_str(col(row, "季节")),
                category=_str(col(row, "类目")),
                product_category=_str(col(row, "商品分类")),
                virtual_category=_str(col(row, "虚拟分类")),
                description=_str(col(row, "款式备注")),
                colors_active=_str(col(row, "在售颜色")),
                colors_discontinued=_str(col(row, "淘汰颜色")),
                color_remark=_str(col(row, "颜色备注")),
                sizes=_str(col(row, "尺码")),
                size_specs=_str(col(row, "号型")),
                size_remark=_str(col(row, "尺码备注")),
                printable_area=_str(col(row, "可印花范围")),
                fabric_composition=_str(col(row, "面料成分")),
                fabric_composition_en=_str(col(row, "Fabric Ingredients")),
                hot_wind_composition=_str(col(row, "热风成分")),
                fabric_name=_str(col(row, "面料名称")),
                fabric_weight=_str(col(row, "面料克重")),
                blank_weight=float(col(row, "白坯重量")) if col(row, "白坯重量") is not None else None,
                dev_date=_excel_date(col(row, "开发时间")),
                tag_price=float(col(row, "吊牌价")) if col(row, "吊牌价") is not None else None,
                premium_tag_price=float(col(row, "高价品牌吊牌价")) if col(row, "高价品牌吊牌价") is not None else None,
                exec_standard=_str(col(row, "执行标准")),
                safety_category=_str(col(row, "安全技术类别")),
                product_type=_str(col(row, "产品分类")),
            )
            existing = crud.get_style_by_code(db, code)
            if existing:
                crud.update_style(db, existing.id, schemas.StyleUpdate(**{k: v for k, v in payload.model_dump().items() if k != "code"}))
                stats["updated"] += 1
            else:
                crud.create_style(db, payload)
                stats["created"] += 1
            count += 1
        except Exception as e:
            errors.append(f"第{row_idx}行处理失败: {e}")
            if len(errors) >= 20:
                break
    return count, errors, stats


def _import_prints(ws, db: Session):
    """支持原始印花.xlsx格式（按列名匹配）和模板格式"""
    headers = [cell.value for cell in ws[1]]
    _require_headers(headers, {"商品编码": ("商品编码",), "图案名称": ("图案名称",)})

    def col(row_vals, *names):
        idx = _header_index(headers, *names)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    count, errors = 0, []
    stats = {"created": 0, "updated": 0, "skipped": 0, "warnings": []}
    seen_codes: dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        code_val = col(row, "商品编码")
        name_val = col(row, "图案名称")
        if not code_val and not name_val:
            stats["skipped"] += 1
            continue
        code = _str(code_val)
        name = _str(name_val)
        if not code:
            errors.append(f"第{row_idx}行：商品编码不能为空")
            continue
        if not name:
            errors.append(f"第{row_idx}行：图案名称不能为空")
            continue
        if code in seen_codes:
            stats["warnings"].append(f"第{row_idx}行：印花编码 {code} 与第{seen_codes[code]}行重复，后者覆盖前者")
        seen_codes[code] = row_idx
        try:
            payload = schemas.PrintCreate(
                code=code,
                name=name,
                pattern_size=_str(col(row, "图案大小")),
                pattern_spec=_str(col(row, "图案规格")),
                craft_attr=_str(col(row, "工艺属性")),
                zwx_style_code=_str(col(row, "真维斯款号")),
                zwx_replace_code=_str(col(row, "真维斯替换编码")),
                zwx_replace_style=_str(col(row, "真维斯替换款号")),
                jwco_style_code=_str(col(row, "JWCO款号")),
                jwco_replace_code=_str(col(row, "JWCO替换编码")),
                jwco_replace_style=_str(col(row, "JWCO替换款号")),
                city_style_code=_str(col(row, "CITY款号")),
                city_replace_code=_str(col(row, "CITY替换编码")),
                city_replace_style=_str(col(row, "CITY替换款号")),
                tangshi_style_code=_str(col(row, "唐狮款号")),
                description=_str(col(row, "备注")),
            )
            existing = crud.get_print_by_code(db, code)
            if existing:
                crud.update_print(db, existing.id, schemas.PrintUpdate(**{k: v for k, v in payload.model_dump().items() if k != "code"}))
                stats["updated"] += 1
            else:
                crud.create_print(db, payload)
                stats["created"] += 1
            count += 1
        except Exception as e:
            errors.append(f"第{row_idx}行处理失败: {e}")
            if len(errors) >= 20:
                break
    return count, errors, stats


def _import_positions(ws, db: Session):
    headers = [cell.value for cell in ws[1]]
    _require_headers(headers, {"位置编号": ("位置编号",), "位置名称": ("位置名称",)})

    def col(row_vals, name):
        idx = _header_index(headers, name)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    count, errors = 0, []
    stats = {"created": 0, "updated": 0, "skipped": 0, "warnings": []}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not col(row, "位置编号") and not col(row, "位置名称"):
            stats["skipped"] += 1
            continue
        code, name = _str(col(row, "位置编号")), _str(col(row, "位置名称"))
        if not code or not name:
            errors.append(f"第{row_idx}行：编号和名称不能为空")
            continue
        payload = schemas.PositionCreate(
            code=code, name=name,
            area=_str(col(row, "区域")),
            description=_str(col(row, "备注")),
        )
        existing = crud.get_position_by_code(db, code)
        if existing:
            if existing.name != name:
                errors.append(f"第{row_idx}行：位置编号 {code} 已绑定名称 '{existing.name}'，不能改为 '{name}'")
                continue
            else:
                stats["skipped"] += 1
                continue
        else:
            existing_name = crud.get_position_by_name(db, name)
            if existing_name:
                errors.append(f"第{row_idx}行：位置名称 '{name}' 已绑定编号 {existing_name.code}，不能改为 {code}")
                continue
            crud.create_position(db, payload)
            stats["created"] += 1
        count += 1
    return count, errors, stats


def _import_positions_dict(data: dict, db: Session):
    CATEGORY_MAP = {
        "big_position": "大图位置",
        "small_position": "小图位置",
        "combination_position": "组合位置",
    }
    if not isinstance(data, dict) or not any(isinstance(items, dict) for items in data.values()):
        raise HTTPException(400, "位置 JSON 必须是分组字典格式")

    count, errors = 0, []
    stats = {"created": 0, "updated": 0, "skipped": 0, "warnings": []}
    seen_codes: dict[str, str] = {}
    seen_names: dict[str, str] = {}
    for group_key, items in data.items():
        area = CATEGORY_MAP.get(group_key, group_key)
        if not isinstance(items, dict):
            continue
        for pos_name, pos_code in items.items():
            code = _str(str(pos_code))
            name = _str(str(pos_name))
            if not code or not name:
                stats["skipped"] += 1
                continue

            if code in seen_codes and seen_codes[code] != name:
                errors.append(f"位置编号 {code} 在文件内同时绑定 '{seen_codes[code]}' 和 '{name}'")
                continue
            if name in seen_names and seen_names[name] != code:
                errors.append(f"位置名称 '{name}' 在文件内同时绑定 {seen_names[name]} 和 {code}")
                continue
            seen_codes[code] = name
            seen_names[name] = code

            try:
                existing = crud.get_position_by_code(db, code)
                if existing:
                    if existing.name != name:
                        errors.append(f"位置编号 {code} 已绑定名称 '{existing.name}'，不能改为 '{name}'")
                    else:
                        stats["skipped"] += 1
                    continue

                existing_name = crud.get_position_by_name(db, name)
                if existing_name:
                    errors.append(f"位置名称 '{name}' 已绑定编号 {existing_name.code}，不能改为 {code}")
                    continue

                crud.create_position(db, schemas.PositionCreate(code=code, name=name, area=area))
                stats["created"] += 1
                count += 1
            except Exception as e:
                errors.append(f"{code} 处理失败: {e}")
    return count, errors, stats


import re as _re

# 导入时跳过的特殊印花（不算真实印花，无需限定）
_SPECIAL_PRINTS = {"纯色", "福袋", "自搭"}


def _split_multi(s: str):
    """按中英文逗号/顿号/换行拆分，去空白"""
    return [x.strip() for x in _re.split(r'[,，、\n\r]+', s) if x.strip()]


def _parse_code_list(raw: str):
    """
    解析编码列表字符串（款式/印花/位置）：
    - 返回 None  → 空值
    - 返回 list  → 编码列表（已去重排序，印花已剔除特殊值）
    """
    if not raw:
        return None
    codes = _split_multi(raw)
    # 如果是印花编码，过滤特殊印花
    codes = [c for c in codes if c not in _SPECIAL_PRINTS]
    return sorted(set(codes)) if codes else None


def _import_restrictions(ws, db: Session):
    """
    特殊款式限定位置表导入（全量替换）

    Excel格式（4列）：款式 | 位置 | 印花 | 限定款式

    支持3种规则类型：
    1. style_ban (类型1): 款式列有值 + 位置列为空 + 印花列为空
    2. position_restriction (类型2): 款式列为空 + 位置列有值 + (印花列有值 或 限定款式列有值)
    3. style_position (类型3): 款式列有值 + 位置列有值

    处理逻辑：
    - 使用 pandas 读取 Excel
    - 读取真实合并单元格区域，仅类型3的款式合并行继承合并区域首行款式
    - 保存原始款式列的空值标记，用于区分类型2和类型3
    - 先分类解析，再按类型1 -> 类型2 -> 类型3强制导入
    - 类型3若与已成功导入的类型2存在相同款式+位置，则跳过
    - 列表字段部分编码不存在时，保留有效编码继续导入，并记录错误
    - 全量替换：先清空所有规则，再导入新数据
    - 自动过滤特殊印花（纯色/自搭/福袋）
    """
    # 1. 提取数据到 pandas DataFrame
    headers = [cell.value for cell in ws[1]]
    _require_headers(headers, {
        "款式": ("款式", "白坯款式编码"),
        "位置": ("位置", "位置名称"),
        "印花": ("印花", "印花（逗号分隔，空=不限）", "允许印花编码(逗号分隔，留空=不限)"),
        "限定款式": ("限定款式", "限定款式（逗号分隔）"),
    })
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    df = pd.DataFrame(data, columns=headers)

    # 2. 标准化列名
    df.columns = df.columns.str.strip()
    col_map = {
        "款式": "style",
        "白坯款式编码": "style",
        "位置": "position",
        "位置名称": "position",
        "印花": "print",
        "印花（逗号分隔，空=不限）": "print",
        "允许印花编码(逗号分隔，留空=不限)": "print",
        "限定款式": "limit_styles",
        "限定款式（逗号分隔）": "limit_styles"
    }

    # 尝试匹配列名（支持带*的列名）
    style_col_idx = None
    for col in df.columns:
        clean_col = col.rstrip("*").strip()
        if clean_col in col_map:
            if col_map[clean_col] == "style":
                style_col_idx = list(df.columns).index(col) + 1
            df.rename(columns={col: col_map[clean_col]}, inplace=True)

    # 确保必要的列存在
    required_cols = ["style", "position", "print"]
    for col in required_cols:
        if col not in df.columns:
            df[col] = None

    if "limit_styles" not in df.columns:
        df["limit_styles"] = None

    # 3. 保存原始款式列的空值标记（用于区分类型2和类型3）
    # 空值包括：None, NaN, 空字符串
    df["_original_style_empty"] = df["style"].isna() | (df["style"].fillna("").astype(str).str.strip() == "")

    # 4. 处理真实合并单元格：只有款式列落在合并区域内的行才继承款式
    merged_style_by_row = {}
    if style_col_idx:
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= style_col_idx <= merged_range.max_col:
                merged_style = ws.cell(merged_range.min_row, style_col_idx).value
                for excel_row in range(merged_range.min_row, merged_range.max_row + 1):
                    if excel_row >= 2:
                        merged_style_by_row[excel_row] = merged_style
    df["_merged_style"] = [merged_style_by_row.get(i + 2) for i in range(len(df))]

    # 5. 先解析分类，再按类型1 -> 类型2 -> 类型3强制导入
    type1_rules = []
    type2_rules = []
    type3_rules = []
    filtered_count = 0  # 被过滤的特殊印花记录数
    skipped_duplicate_count = 0  # 被类型2覆盖或自身重复的类型3记录数
    errors = []

    def _append_missing_error(row_num: int, label: str, missing_codes: list[str]):
        if not missing_codes:
            return
        preview = "、".join(missing_codes[:5])
        suffix = f" 等{len(missing_codes)}个" if len(missing_codes) > 5 else ""
        errors.append(f"第{row_num}行：{label}不存在：{preview}{suffix}")

    def _validate_existing_prints(row_num: int, print_codes: list[str] | None):
        if not print_codes:
            return None
        missing = []
        for code in print_codes:
            if crud.get_print_by_code(db, code):
                continue
            else:
                missing.append(code)
        _append_missing_error(row_num, "印花编码", missing)
        return print_codes if not missing else None

    def _validate_existing_styles(row_num: int, style_codes: list[str] | None):
        if not style_codes:
            return None
        missing = []
        for code in style_codes:
            if crud.get_style_by_code(db, code):
                continue
            else:
                missing.append(code)
        _append_missing_error(row_num, "款式编码", missing)
        return style_codes if not missing else None

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel 行号（从2开始）

        # 提取并清理数据
        style_code = _str(row.get("style"))
        position_name = _str(row.get("position"))
        print_codes_str = _str(row.get("print"))
        limit_styles_str = _str(row.get("limit_styles"))
        original_style_empty = row.get("_original_style_empty", False)
        merged_style_code = _str(row.get("_merged_style"))
        is_style_merged_row = bool(merged_style_code)
        effective_style_code = merged_style_code if original_style_empty and is_style_merged_row else style_code

        # 跳过完全空行
        if not any([effective_style_code, position_name, print_codes_str, limit_styles_str]):
            continue

        # 判断规则类型并分类
        try:
            # 类型1: style_ban（原始款式列有值 + 位置为空 + 印花为空）
            if not original_style_empty and effective_style_code and not position_name and not print_codes_str:
                if not _validate_existing_styles(row_num, [effective_style_code]):
                    continue
                type1_rules.append((row_num, schemas.StylePositionRuleCreate(
                    rule_type=1,
                    position_code=None,
                    style_codes=[effective_style_code],
                    print_codes=None,
                    is_active=True
                )))

            # 类型2: position_restriction（原始款式列为空，且不在款式合并区域内 + 位置有值）
            elif original_style_empty and not is_style_merged_row and position_name:
                print_codes = _parse_code_list(print_codes_str)
                limit_style_codes = _parse_code_list(limit_styles_str)

                # 印花全是特殊印花时，若没有限定款式则整行无需生成规则。
                if print_codes_str and not print_codes and not limit_style_codes:
                    filtered_count += 1
                    continue

                # 类型2要求至少有一个维度有值
                if not print_codes and not limit_style_codes:
                    errors.append(f"第{row_num}行：位置限定规则的印花和限定款式不能同时为空")
                    continue

                # 通过位置名称查找位置编码
                position = crud.get_position_by_name(db, position_name)
                if not position:
                    errors.append(f"第{row_num}行：位置 '{position_name}' 不存在")
                    continue

                print_codes = _validate_existing_prints(row_num, print_codes)
                limit_style_codes = _validate_existing_styles(row_num, limit_style_codes)
                if errors and any(error.startswith(f"第{row_num}行：") for error in errors):
                    continue
                if not print_codes and not limit_style_codes:
                    errors.append(f"第{row_num}行：位置限定规则过滤无效编码后为空")
                    continue

                type2_rules.append((row_num, schemas.StylePositionRuleCreate(
                    rule_type=2,
                    position_code=position.code,  # 使用位置编码
                    style_codes=limit_style_codes,
                    print_codes=print_codes,
                    is_active=True
                )))

            # 类型3: style_position（原始款式列有值或处于款式合并区域内 + 位置有值）
            elif effective_style_code and (not original_style_empty or is_style_merged_row) and position_name:
                if not effective_style_code:
                    errors.append(f"第{row_num}行：款式位置规则缺少款式编码")
                    continue
                if not _validate_existing_styles(row_num, [effective_style_code]):
                    continue

                # 通过位置名称查找位置编码
                position = crud.get_position_by_name(db, position_name)
                if not position:
                    errors.append(f"第{row_num}行：位置 '{position_name}' 不存在")
                    continue

                print_codes = _parse_code_list(print_codes_str)

                # 检查是否被过滤（印花全是特殊印花）
                if print_codes_str and not print_codes:
                    filtered_count += 1
                    continue

                print_codes = _validate_existing_prints(row_num, print_codes)
                if print_codes_str and any(error.startswith(f"第{row_num}行：印花编码不存在") for error in errors):
                    continue

                type3_rules.append((row_num, schemas.StylePositionRuleCreate(
                    rule_type=3,
                    position_code=position.code,  # 使用位置编码
                    style_codes=[effective_style_code],
                    print_codes=print_codes,  # None 表示不限印花
                    is_active=True
                )))

            else:
                # 无法识别的规则类型
                errors.append(f"第{row_num}行：无法识别规则类型（款式={style_code}, 位置={position_name}）")

        except Exception as e:
            errors.append(f"第{row_num}行处理失败: {e}")
            continue

    # 6. 有解析/引用错误时不替换旧规则
    if errors:
        return 0, errors, filtered_count, skipped_duplicate_count

    # 7. 清空现有规则（全量替换）
    db.query(models.StylePositionRule).delete()
    db.commit()

    # 8. 按类型1 -> 类型2 -> 类型3导入，类型2覆盖同款式+位置的类型3
    type1_count, type2_count, type3_count = 0, 0, 0
    seen_type1_keys = set()
    seen_type2_keys = set()
    type2_style_position_keys = set()
    seen_type3_keys = set()

    for row_num, rule_data in type1_rules:
        type1_key = tuple(rule_data.style_codes or [])
        if type1_key in seen_type1_keys:
            skipped_duplicate_count += 1
            continue
        seen_type1_keys.add(type1_key)
        try:
            crud.create_style_position_rule(db, rule_data)
            type1_count += 1
        except Exception as e:
            errors.append(f"第{row_num}行处理失败: {e}")

    for row_num, rule_data in type2_rules:
        type2_key = (
            rule_data.position_code,
            tuple(rule_data.style_codes or []),
            tuple(rule_data.print_codes or []),
        )
        if type2_key in seen_type2_keys:
            skipped_duplicate_count += 1
            continue
        seen_type2_keys.add(type2_key)
        try:
            crud.create_style_position_rule(db, rule_data)
            type2_count += 1
            for style_code in rule_data.style_codes or []:
                type2_style_position_keys.add((style_code, rule_data.position_code))
        except Exception as e:
            errors.append(f"第{row_num}行处理失败: {e}")

    for row_num, rule_data in type3_rules:
        style_code = rule_data.style_codes[0] if rule_data.style_codes else None
        if (style_code, rule_data.position_code) in type2_style_position_keys:
            skipped_duplicate_count += 1
            continue

        type3_key = (
            style_code,
            rule_data.position_code,
            tuple(rule_data.print_codes or []),
        )
        if type3_key in seen_type3_keys:
            skipped_duplicate_count += 1
            continue
        seen_type3_keys.add(type3_key)

        try:
            crud.create_style_position_rule(db, rule_data)
            type3_count += 1
        except Exception as e:
            errors.append(f"第{row_num}行处理失败: {e}")

    total_count = type1_count + type2_count + type3_count
    return total_count, errors, filtered_count, skipped_duplicate_count


def _result(label: str, count: int, errors: list, stats: dict | None = None) -> schemas.ImportResult:
    msg = f"导入完成：{label} {count} 条"
    stats = stats or {}
    if stats:
        parts = []
        if stats.get("created"):
            parts.append(f"新增 {stats['created']} 条")
        if stats.get("updated"):
            parts.append(f"更新 {stats['updated']} 条")
        if stats.get("skipped"):
            parts.append(f"跳过 {stats['skipped']} 条")
        if parts:
            msg += "（" + "，".join(parts) + "）"
    warnings = stats.get("warnings", []) if stats else []
    if warnings:
        msg += f"；有 {len(warnings)} 条提示"
    if errors:
        msg += f"；有 {len(errors)} 条错误"
    return schemas.ImportResult(
        success=len(errors) == 0,
        message=msg,
        details={
            "counts": {label: count, **{k: v for k, v in stats.items() if k != "warnings"}},
            "errors": errors[:20],
            "warnings": warnings[:20],
        },
    )


def _first_sheet(wb, filename: str):
    """取工作簿第一个 sheet，若无则报错"""
    if not wb.sheetnames:
        raise HTTPException(400, "文件中没有找到任何 Sheet")
    return wb.active


# ── 分模块导入接口 ─────────────────────────────────────

@router.post("/import/styles", summary="导入款式", response_model=schemas.ImportResult)
async def import_styles(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 文件")
    wb = _parse_wb(await file.read(), file.filename)
    ws = wb["款式"] if "款式" in wb.sheetnames else _first_sheet(wb, file.filename)
    count, errors, stats = _import_styles(ws, db)
    return _result("款式", count, errors, stats)


@router.post("/import/prints", summary="导入印花", response_model=schemas.ImportResult)
async def import_prints(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 文件")
    wb = _parse_wb(await file.read(), file.filename)
    ws = wb["印花"] if "印花" in wb.sheetnames else _first_sheet(wb, file.filename)
    count, errors, stats = _import_prints(ws, db)
    return _result("印花", count, errors, stats)


@router.post("/import/positions", summary="导入位置（JSON/txt）", response_model=schemas.ImportResult)
async def import_positions(file: UploadFile = File(...), db: Session = Depends(get_db)):
    fname = file.filename or ""
    if not fname.endswith((".json", ".txt")):
        raise HTTPException(400, "请上传 JSON/txt 格式的贴图位置字典文件")
    content = await file.read()
    try:
        data = json.loads(content.decode("utf-8"))
    except Exception as e:
        raise HTTPException(400, f"JSON 解析失败: {e}")

    count, errors, stats = _import_positions_dict(data, db)
    return _result("位置", count, errors, stats)


@router.post("/import/restrictions", summary="导入限定", response_model=schemas.ImportResult)
async def import_restrictions(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 文件")
    wb = _parse_wb(await file.read(), file.filename)
    ws = _first_sheet(wb, file.filename)
    count, errors, filtered_count, skipped_duplicate_count = _import_restrictions(ws, db)

    # 构造返回消息
    msg = f"导入完成：限定 {count} 条"
    if filtered_count > 0:
        msg += f"；已过滤 {filtered_count} 条特殊印花记录（纯色/自搭/福袋）"
    if skipped_duplicate_count > 0:
        msg += f"；已跳过 {skipped_duplicate_count} 条重复类型3记录"
    if errors:
        msg += f"；有 {len(errors)} 条错误"

    return schemas.ImportResult(
        success=len(errors) == 0,
        message=msg,
        details={
            "counts": {
                "限定": count,
                "filtered": filtered_count,
                "skipped_duplicate": skipped_duplicate_count,
            },
            "errors": errors[:5],
        },
    )


# ── 分模块模版下载 ─────────────────────────────────────

@router.get("/template/styles", summary="下载款式导入模板")
def template_styles(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "款式"
    ws.append(STYLE_HEADERS)
    _apply_header_style(ws)
    samples = crud.get_styles(db, page_size=3)["items"]
    for s in samples:
        ws.append([
            s.brand_attr, s.attr, s.fabric_type, s.year, s.gender, s.season,
            s.category, s.product_category, s.virtual_category,
            s.product_code, s.code, s.description,
            s.colors_active, s.colors_discontinued, s.color_remark,
            s.sizes, s.size_specs, s.size_remark,
            s.printable_area,
            s.fabric_composition, s.fabric_composition_en, s.hot_wind_composition,
            s.fabric_name, s.fabric_weight, s.blank_weight,
            s.dev_date, s.tag_price, s.premium_tag_price,
            s.exec_standard, s.safety_category, s.product_type,
        ])
    return _stream_wb(wb, "template_styles.xlsx")


@router.get("/template/prints", summary="下载印花导入模板")
def template_prints(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "印花"
    ws.append(PRINT_HEADERS)
    _apply_header_style(ws)
    samples = crud.get_prints(db, page_size=3)["items"]
    for p in samples:
        ws.append([
            p.name, p.pattern_size, p.pattern_spec, p.craft_attr, p.code,
            p.zwx_style_code, p.zwx_replace_code, p.zwx_replace_style,
            p.jwco_style_code, p.jwco_replace_code, p.jwco_replace_style,
            p.city_style_code, p.city_replace_code, p.city_replace_style,
            p.tangshi_style_code, p.description,
        ])
    return _stream_wb(wb, "template_prints.xlsx")


@router.get("/template/positions", summary="下载位置模板（JSON/txt）")
def template_positions(db: Session = Depends(get_db)):
    """返回与导入格式完全一致的 JSON/txt 文件，取库中实际数据"""
    AREA_KEY = {
        "大图位置": "big_position",
        "小图位置": "small_position",
        "组合位置": "combination_position",
    }
    groups: dict = {}
    for p in crud.get_positions(db, page_size=99999)["items"]:
        key = AREA_KEY.get(p.area or "", p.area or "other")
        groups.setdefault(key, {})[p.name] = p.code
    content = json.dumps(groups, ensure_ascii=False, indent=2).encode("utf-8")
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=template_positions.txt"},
    )


@router.get("/template/restrictions", summary="下载限定导入模板")
def template_restrictions(db: Session = Depends(get_db)):
    """
    下载限定导入模板（4列格式）
    支持3种规则类型：
    - 类型1 (style_ban): 款式有值，位置为空，印花为空
    - 类型2 (position_restriction): 款式为空，位置有值，印花或限定款式有值
    - 类型3 (style_position): 款式有值，位置有值，印花可选
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "限定"
    ws.append(["款式*", "位置", "印花（逗号分隔，空=不限）", "限定款式（逗号分隔）"])
    _apply_header_style(ws)

    # 添加示例数据
    rules = crud.get_style_position_rules(db, page_size=10)["items"]

    for r in rules:
        # 根据规则类型生成不同的行
        if r.rule_type == 1:  # style_ban
            # 款式全禁：款式有值，其他为空
            style_codes = r.style_ids.split(",") if r.style_ids else []
            for style_id in style_codes:
                style = crud.get_style(db, int(style_id))
                if style:
                    ws.append([style.code, "", "", ""])

        elif r.rule_type == 2:  # position_restriction
            # 位置限定：款式为空，位置有值，印花和限定款式有值
            position_name = r.position.name if r.position else ""

            # 获取印花编码
            print_codes = []
            if r.print_ids:
                for print_id in r.print_ids.split(","):
                    print_obj = crud.get_print(db, int(print_id))
                    if print_obj:
                        print_codes.append(print_obj.code)
            print_codes_str = ",".join(print_codes) if print_codes else ""

            # 获取款式编码
            style_codes = []
            if r.style_ids:
                for style_id in r.style_ids.split(","):
                    style = crud.get_style(db, int(style_id))
                    if style:
                        style_codes.append(style.code)
            style_codes_str = ",".join(style_codes) if style_codes else ""

            ws.append(["", position_name, print_codes_str, style_codes_str])

        elif r.rule_type == 3:  # style_position
            # 款式位置限定：款式有值，位置有值，印花可选
            style_codes = r.style_ids.split(",") if r.style_ids else []
            position_name = r.position.name if r.position else ""

            # 获取印花编码
            print_codes = []
            if r.print_ids:
                for print_id in r.print_ids.split(","):
                    print_obj = crud.get_print(db, int(print_id))
                    if print_obj:
                        print_codes.append(print_obj.code)
            print_codes_str = ",".join(print_codes) if print_codes else ""

            for style_id in style_codes:
                style = crud.get_style(db, int(style_id))
                if style:
                    ws.append([style.code, position_name, print_codes_str, ""])

    return _stream_wb(wb, "template_restrictions.xlsx")


# ── 导出数据 ──────────────────────────────────────────────
# entities 参数：逗号分隔，支持 styles / prints / positions / rules / bans
# 不传或传 all 时导出全部

_ALL_ENTITIES = ["styles", "prints", "positions", "rules", "bans"]


@router.get("/export", summary="导出数据为Excel（支持选择实体）")
def export_excel(
    entities: Optional[str] = Query(None, description="逗号分隔: styles,prints,positions,rules,bans；不传=全量"),
    db: Session = Depends(get_db),
):
    selected = (
        {e.strip() for e in entities.split(",") if e.strip()}
        if entities and entities.strip() and entities.strip() != "all"
        else set(_ALL_ENTITIES)
    )

    wb = Workbook()
    wb.remove(wb.active)

    if "styles" in selected:
        ws = wb.create_sheet("款式")
        ws.append(STYLE_HEADERS)
        _apply_header_style(ws)
        for s in crud.get_styles(db, page_size=99999)["items"]:
            ws.append([
                s.brand_attr, s.attr, s.fabric_type, s.year, s.gender, s.season,
                s.category, s.product_category, s.virtual_category,
                s.product_code, s.code, s.description,
                s.colors_active, s.colors_discontinued, s.color_remark,
                s.sizes, s.size_specs, s.size_remark,
                s.printable_area,
                s.fabric_composition, s.fabric_composition_en, s.hot_wind_composition,
                s.fabric_name, s.fabric_weight, s.blank_weight,
                s.dev_date, s.tag_price, s.premium_tag_price,
                s.exec_standard, s.safety_category, s.product_type,
            ])

    if "prints" in selected:
        ws = wb.create_sheet("印花")
        ws.append(PRINT_HEADERS)
        _apply_header_style(ws)
        for p in crud.get_prints(db, page_size=99999)["items"]:
            ws.append([
                p.name, p.pattern_size, p.pattern_spec, p.craft_attr, p.code,
                p.zwx_style_code, p.zwx_replace_code, p.zwx_replace_style,
                p.jwco_style_code, p.jwco_replace_code, p.jwco_replace_style,
                p.city_style_code, p.city_replace_code, p.city_replace_style,
                p.tangshi_style_code, p.description,
            ])

    if "positions" in selected:
        ws = wb.create_sheet("位置")
        ws.append(POSITION_HEADERS)
        _apply_header_style(ws)
        for pos in crud.get_positions(db, page_size=99999)["items"]:
            ws.append([pos.code, pos.name, pos.area, pos.description])

    if "rules" in selected:
        ws = wb.create_sheet("限定规则")
        ws.append(["款式", "位置", "印花（逗号分隔）", "限定款式（逗号分隔）", "规则类型", "是否启用"])
        _apply_header_style(ws)
        for r in crud.get_style_position_rules(db, page_size=999999)["items"]:
            # 根据规则类型生成不同的行
            rule_type_name = {1: "款式全禁", 2: "位置限定", 3: "款式位置限定"}.get(r.rule_type, "未知")

            if r.rule_type == 1:  # style_ban
                style_codes = r.style_ids.split(",") if r.style_ids else []
                for style_id in style_codes:
                    style = crud.get_style(db, int(style_id))
                    if style:
                        ws.append([
                            style.code,
                            "",
                            "",
                            "",
                            rule_type_name,
                            1 if r.is_active else 0
                        ])

            elif r.rule_type == 2:  # position_restriction
                position_name = r.position.name if r.position else ""

                # 获取印花编码
                print_codes = []
                if r.print_ids:
                    for print_id in r.print_ids.split(","):
                        print_obj = crud.get_print(db, int(print_id))
                        if print_obj:
                            print_codes.append(print_obj.code)
                print_codes_str = ",".join(print_codes) if print_codes else ""

                # 获取款式编码
                style_codes = []
                if r.style_ids:
                    for style_id in r.style_ids.split(","):
                        style = crud.get_style(db, int(style_id))
                        if style:
                            style_codes.append(style.code)
                style_codes_str = ",".join(style_codes) if style_codes else ""

                ws.append([
                    "",
                    position_name,
                    print_codes_str,
                    style_codes_str,
                    rule_type_name,
                    1 if r.is_active else 0
                ])

            elif r.rule_type == 3:  # style_position
                style_codes = r.style_ids.split(",") if r.style_ids else []
                position_name = r.position.name if r.position else ""

                # 获取印花编码
                print_codes = []
                if r.print_ids:
                    for print_id in r.print_ids.split(","):
                        print_obj = crud.get_print(db, int(print_id))
                        if print_obj:
                            print_codes.append(print_obj.code)
                print_codes_str = ",".join(print_codes) if print_codes else ""

                for style_id in style_codes:
                    style = crud.get_style(db, int(style_id))
                    if style:
                        ws.append([
                            style.code,
                            position_name,
                            print_codes_str,
                            "",
                            rule_type_name,
                            1 if r.is_active else 0
                        ])

    if "bans" in selected:
        # bans 已经包含在 rules 中（类型1），这里保留是为了向后兼容
        ws = wb.create_sheet("全禁款式")
        ws.append(["白坯款式编码", "是否启用"])
        _apply_header_style(ws)
        for r in crud.get_style_position_rules(db, page_size=999999)["items"]:
            if r.rule_type == 1:  # style_ban
                style_codes = r.style_ids.split(",") if r.style_ids else []
                for style_id in style_codes:
                    style = crud.get_style(db, int(style_id))
                    if style:
                        ws.append([style.code, 1 if r.is_active else 0])

    if not wb.sheetnames:
        raise HTTPException(400, "未选择任何有效的导出实体")

    name = f"export_{selected.pop()}.xlsx" if len(selected) == 1 else "export_all.xlsx"
    return _stream_wb(wb, name)
