"""
测试限定表导入逻辑

使用方法：
1. 准备一个测试 Excel 文件（4列：款式、位置、印花、限定款式）
2. 运行：uv run python test_import_restrictions.py <excel_file_path>
"""

import sys
from openpyxl import load_workbook
from app.database import SessionLocal
from app.routers.excel_io import _import_restrictions


def test_import(excel_path: str):
    """测试导入限定表"""
    print(f"正在读取 Excel 文件: {excel_path}")

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        print(f"工作表名称: {ws.title}")
        print(f"总行数: {ws.max_row}")

        # 显示前10行数据
        print("\n前10行数据预览：")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            print(f"第{i}行: {row}")

        # 执行导入
        print("\n开始导入...")
        db = SessionLocal()
        try:
            count, errors, filtered_count, skipped_duplicate_count, warnings = _import_restrictions(ws, db)
            print(f"\n导入完成！")
            print(f"成功导入: {count} 条规则")
            print(f"过滤特殊印花: {filtered_count} 条")
            print(f"跳过重复类型3: {skipped_duplicate_count} 条")

            if errors:
                print(f"\n错误信息 ({len(errors)} 条):")
                for error in errors[:5]:  # 只显示前5条错误
                    print(f"  - {error}")
                if len(errors) > 5:
                    print(f"  ... 还有 {len(errors) - 5} 条错误")
            else:
                print("\n没有错误！")

            if warnings:
                print(f"\n提示信息 ({len(warnings)} 条):")
                for warning in warnings[:5]:
                    print(f"  - {warning}")
                if len(warnings) > 5:
                    print(f"  ... 还有 {len(warnings) - 5} 条提示")

        finally:
            db.close()

    except Exception as e:
        print(f"\n错误: {e}")
        from loguru import logger
        logger.exception(e)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: uv run python test_import_restrictions.py <excel_file_path>")
        sys.exit(1)

    test_import(sys.argv[1])
