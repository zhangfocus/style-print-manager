"""
生成测试用的限定表 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "限定"

# 添加表头
headers = ["款式", "位置", "印花", "限定款式"]
ws.append(headers)

# 设置表头样式
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# 添加测试数据
test_data = [
    # 类型3: 款式位置限定（款式有值 + 位置有值）
    ["款式A", "前胸", "印花1,印花2,印花3", ""],
    ["款式A", "后背", "印花4,印花5", ""],
    ["款式A", "袖口", "", ""],  # 不限印花

    # 类型1: 款式全禁（款式有值 + 位置为空 + 印花为空）
    ["款式B", "", "", ""],

    # 类型3: 另一个款式的位置限定
    ["款式C", "前胸", "印花1", ""],
    ["款式C", "领口", "印花2,印花3", ""],

    # 类型2: 位置限定（款式为空 + 位置有值 + 印花或限定款式有值）
    ["", "特殊位置1", "印花6,印花7", "款式D,款式E"],
    ["", "特殊位置2", "", "款式F,款式G"],  # 不限印花，只限款式
    ["", "特殊位置3", "印花8", ""],  # 不限款式，只限印花
]

for row in test_data:
    ws.append(row)

# 保存文件
filename = "test_restrictions.xlsx"
wb.save(filename)
print(f"测试文件已生成: {filename}")
print("\n文件内容：")
print("=" * 80)
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f"第{i}行: {row}")
